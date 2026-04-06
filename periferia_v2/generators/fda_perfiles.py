"""
generators/fda_perfiles.py
Responsable: Heidy

Edita slides Perfiles y FDA del PPTX de la filial.

Lógica Perfiles:
  - Fuente de verdad: hoja 'Perfiles' del Excel Generales_para_todos.xlsx
  - Se traen TODOS los perfiles de CADA torre seleccionada (no solo uno)
  - Se paginan en slides de máximo 4 perfiles
  - Slides con menos de 4 perfiles se centran horizontalmente
  - Los slots vacíos se ELIMINAN del XML (no solo se ocultan)
  - Los slides se localizan por contenido, no por índice fijo

Lógica FDA:
  - Si hay 1 torre → mostrar ítems de esa torre (max 6)
  - Si hay más de 1 torre → mostrar cláusula general aplicable a todas
  - Si la torre es QA → ocultar card verde de QA (ya está como torre)
"""

import copy, io, re, zipfile, unicodedata
from pathlib import Path
from lxml import etree
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / 'data'
GENERALES = DATA_DIR / 'Generales_para_todos.xlsx'

A  = 'http://schemas.openxmlformats.org/drawingml/2006/main'
P  = 'http://schemas.openxmlformats.org/presentationml/2006/main'
R  = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

SLIDE_WIDTH_EMU = 9144000  # 10 pulgadas en EMU (estándar widescreen)

# ── Identificadores de slides por contenido (resistente a duplicación) ───────
# En vez de índice fijo (se rompe al duplicar), buscamos un shape único.
PERFILES_MARKER = 'CuadroTexto 10'   # Shape exclusivo del slide de Perfiles
FDA_MARKER      = 'Rectángulo 10'    # Shape exclusivo del slide de FDA

# ── Slot definitions slide Perfiles ──────────────────────────────────────────
# Cada slot es un grpSp externo en el spTree. Dentro de cada grupo hay shapes
# con nombres únicos por slot. Identificamos el slot por el nombre del shape
# de rol (CuadroTexto N).
#
# Posiciones originales en el template (EMU):
#   Slot 0: off.x=445717   (Grupo 23)
#   Slot 1: off.x=2435455  (Grupo 24)
#   Slot 2: off.x=4539436  (Grupo 32)
#   Slot 3: off.x=6680393  (Grupo 49)
_PERFIL_SLOT_NAMES = [
    ('CuadroTexto 10', 'CuadroTexto 22'),   # Slot 0: (rol, desc)
    ('CuadroTexto 30', 'CuadroTexto 28'),   # Slot 1
    ('CuadroTexto 47', 'CuadroTexto 34'),   # Slot 2
    ('CuadroTexto 53', 'CuadroTexto 51'),   # Slot 3
]
_PERFIL_ROLE_NAMES = {name for name, _ in _PERFIL_SLOT_NAMES}  # set para búsqueda rápida
_PERFIL_CARD_WIDTH = 1550783   # ext.cx de cada grupo (EMU)
_PERFIL_PITCH      = 2078225   # espaciado promedio entre grupos (EMU)

# ── Shapes slide FDA ─────────────────────────────────────────────────────────
BULLET_RECTS = ['Rectángulo 10', 'Rectángulo 13', 'Rectángulo 19',
                'Rectángulo 22', 'Rectángulo 25', 'Rectángulo 28']
QA_CARD        = 'CuadroTexto 32'
TITULO_TECNICO = 'Fuera del Alcance Técnico'

# ── Cláusula general para múltiples torres ───────────────────────────────────
CLAUSULA_GENERAL = [
    'El servicio será ejecutado conforme al alcance técnico aprobado y a la información disponible al momento de la estimación.',
    'Cualquier modificación posterior en requerimientos funcionales, técnicos o de negocio será gestionada mediante la metodología formal de control de cambios.',
    'Cualquier actividad no descrita explícitamente en el alcance aprobado se considerará fuera de alcance.',
    'No incluye actividades de soporte continuo posterior al periodo de garantía definido.',
    'No incluye licenciamiento de herramientas, plataformas o componentes de terceros.',
    'No incluye infraestructura productiva ni ambientes no definidos en el alcance.',
]

# Texto que se muestra (en negrita) cuando no hay descripción en el catálogo
_NO_CATALOG_DESC = 'No encontramos este perfil en la base de datos'

# Máx. caracteres de descripción por tarjeta de perfil (evita desbordamiento)
DESC_MAX_CHARS = 200


# ═══════════════════════════════ utilidades ══════════════════════════════════

def _norm(s):
    s = (s or '').strip().upper()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', s).strip()


def _esc(t):
    return (t or '').replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


def _even_chunks(items, max_per=6):
    """
    Divide items en chunks distribuidos uniformemente (máx max_per por chunk).
    Ej: 7 → [4, 3]  |  13 → [5, 4, 4]  |  6 → [6]  |  1 → [1]
    """
    total = len(items)
    if not total:
        return [[]]
    n = -(-total // max_per)   # ceil(total / max_per)
    base, rem = divmod(total, n)
    chunks, idx = [], 0
    for i in range(n):
        size = base + (1 if i < rem else 0)
        chunks.append(items[idx: idx + size])
        idx += size
    return chunks


def _set_bullet_shapes(root, items):
    """
    Rellena los grupos bullet del slide FDA con los ítems dados.
    - Menos de 6 ítems: elimina los grupos sobrantes y redistribuye los visibles
      uniformemente dentro del bounding box original.
    - Más de 6 ítems: clona grupos extra y los redistribuye dentro del mismo bbox.
    """
    # Mapear BULLET_RECT name → grpSp que lo contiene
    name_to_grp = {}
    for sp in root.iter(f'{{{P}}}sp'):
        nvpr = sp.find(f'.//{{{P}}}cNvPr')
        if nvpr is None:
            continue
        n = nvpr.attrib.get('name', '')
        if n in BULLET_RECTS:
            parent = sp.getparent()
            if parent is not None and parent.tag == f'{{{P}}}grpSp':
                name_to_grp[n] = parent

    base_grps = [name_to_grp[n] for n in BULLET_RECTS if n in name_to_grp]
    if not base_grps:
        return

    n_items = len(items)
    n_base  = len(base_grps)

    # ── Helpers ──────────────────────────────────────────────────────────────
    def _grp_y(grp):
        grpSpPr = grp.find(f'{{{P}}}grpSpPr')
        if grpSpPr is None:
            return None
        xfrm = grpSpPr.find(f'{{{A}}}xfrm')
        if xfrm is None:
            return None
        off = xfrm.find(f'{{{A}}}off')
        return int(off.attrib.get('y', 0)) if off is not None else None

    def _set_grp_y(grp, new_y):
        grpSpPr = grp.find(f'{{{P}}}grpSpPr')
        if grpSpPr is None:
            return
        xfrm = grpSpPr.find(f'{{{A}}}xfrm')
        if xfrm is None:
            return
        off = xfrm.find(f'{{{A}}}off')
        if off is not None:
            off.attrib['y'] = str(new_y)

    def _grp_cy(grp):
        grpSpPr = grp.find(f'{{{P}}}grpSpPr')
        if grpSpPr is None:
            return None
        xfrm = grpSpPr.find(f'{{{A}}}xfrm')
        if xfrm is None:
            return None
        ext = xfrm.find(f'{{{A}}}ext')
        return int(ext.attrib.get('cy', 0)) if ext is not None else None

    def _fill_grp(grp, text):
        for sp in grp:
            if sp.tag != f'{{{P}}}sp':
                continue
            nvpr = sp.find(f'.//{{{P}}}cNvPr')
            if nvpr is not None and nvpr.attrib.get('name') in BULLET_RECTS:
                txb = sp.find(f'{{{P}}}txBody')
                if txb is not None:
                    all_t = txb.findall(f'.//{{{A}}}t')
                    if all_t:
                        all_t[0].text = text
                        for t in all_t[1:]:
                            t.text = ''

    def _remove_grp(grp):
        parent = grp.getparent()
        if parent is not None:
            parent.remove(grp)

    # ── Calcular bbox original (usando los 6 grupos del template) ────────────
    ys = [_grp_y(g) for g in base_grps]
    if any(y is None for y in ys):
        # Fallback sin redistribución
        for i, grp in enumerate(base_grps):
            if i < n_items:
                _fill_grp(grp, items[i])
            else:
                _remove_grp(grp)
        return

    gaps    = [ys[i+1] - ys[i] for i in range(len(ys) - 1)]
    avg_gap = sum(gaps) // len(gaps) if gaps else 400000
    group_cy   = _grp_cy(base_grps[0]) or avg_gap
    y_top      = ys[0]
    total_span = (ys[-1] + group_cy) - y_top   # altura total del bbox original

    # ── Caso: menos o igual de grupos base ───────────────────────────────────
    if n_items <= n_base:
        # Eliminar grupos sobrantes
        for grp in base_grps[n_items:]:
            _remove_grp(grp)

        visible = base_grps[:n_items]
        # Redistribuir los visibles uniformemente en el bbox completo
        if len(visible) > 1:
            step = (total_span - group_cy) / (len(visible) - 1)
            for i, grp in enumerate(visible):
                _set_grp_y(grp, y_top + int(i * step))

        for i, grp in enumerate(visible):
            _fill_grp(grp, items[i])
        return

    # ── Caso: más grupos de lo que hay en el template (clonar) ───────────────
    all_grps = list(base_grps)
    next_id  = 9000

    for i in range(n_items - n_base):
        new_grp = copy.deepcopy(base_grps[0])
        for el in new_grp.iter(f'{{{P}}}cNvPr'):
            el.attrib['id'] = str(next_id)
            next_id += 1
        all_grps[-1].addnext(new_grp)
        all_grps.append(new_grp)

    # Redistribuir uniformemente dentro del bbox original
    N    = len(all_grps)
    step = (total_span - group_cy) / (N - 1) if N > 1 else 0
    for i, grp in enumerate(all_grps):
        _set_grp_y(grp, y_top + int(i * step))
        _fill_grp(grp, items[i])


def _find_desc_in_catalog(rol, perf_db):
    """
    Busca la descripción más cercana para un rol en el catálogo de perfiles.
    Prioridad: coincidencia exacta (normalizada) → coincidencia parcial → ''.
    """
    rol_norm = _norm(rol)
    best_desc = ''
    for torre_perfiles in perf_db.values():
        for p in torre_perfiles:
            p_norm = _norm(p['rol'])
            if p_norm == rol_norm:
                return p['desc']
            if not best_desc and (rol_norm in p_norm or p_norm in rol_norm):
                best_desc = p['desc']
    return best_desc


def _split_fda(cell_value):
    """Divide una celda de Excel en ítems FDA individuales usando '.' como separador."""
    raw = str(cell_value).strip()
    parts = [p.strip() for p in raw.split('.')]
    return [p + '.' for p in parts if p]


def _complement_perfiles(base_perfiles, torres_activas, perf_db):
    """
    Completa base_perfiles con genéricos del catálogo:
    - Si el último slide tiene espacio libre (< 4): lo rellena.
    - Si todos los slides están llenos: agrega un slide extra con genéricos.
    """
    SLOT_SIZE = 4
    existing_rols = {_norm(p['rol']) for p in base_perfiles}

    genericos = []
    for torre in torres_activas:
        key = _norm(torre)
        torre_generics = perf_db.get(key, [])
        if not torre_generics:
            for k in perf_db:
                if key in k or k in key:
                    torre_generics = perf_db[k]
                    break
        for g in torre_generics:
            if _norm(g['rol']) not in existing_rols:
                genericos.append(g)
                existing_rols.add(_norm(g['rol']))

    if not genericos:
        return base_perfiles

    last_slot = len(base_perfiles) % SLOT_SIZE
    if last_slot == 0:
        # Todos los slides llenos → slide extra con genéricos
        return base_perfiles + genericos[:SLOT_SIZE]
    else:
        # Último slide con espacio → completarlo
        needed = SLOT_SIZE - last_slot
        return base_perfiles + genericos[:needed]


def _load_generales():
    """Carga el Excel de generales y retorna dicts por hoja."""
    if not GENERALES.exists():
        return {}, {}
    wb = load_workbook(GENERALES)

    # FDA por torre
    fda_db = {}
    ws = wb['Fuera del Alcance']
    torre_actual = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]).strip() and str(row[0]).strip() != 'Torre':
            torre_actual = _norm(str(row[0]).strip())
            if row[1]:
                fda_db.setdefault(torre_actual, []).extend(_split_fda(row[1]))
        elif row[0] is None and row[1] and torre_actual:
            fda_db.setdefault(torre_actual, []).extend(_split_fda(row[1]))

    # Perfiles por torre
    perf_db = {}
    ws2 = wb['Perfiles']
    torre_actual = None
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]).strip():
            torre_actual = _norm(str(row[0]).strip().replace('TORRE ', ''))
        if torre_actual and row[2] and row[3]:
            perf_db.setdefault(torre_actual, []).append({
                'rol': str(row[2]).strip(),
                'desc': str(row[3]).strip()
            })

    return fda_db, perf_db


def _get_slide_order(pptx_bytes):
    """Retorna lista ordenada de paths de slides según presentation.xml."""
    with zipfile.ZipFile(io.BytesIO(pptx_bytes)) as z:
        rels = etree.fromstring(z.read('ppt/_rels/presentation.xml.rels'))
        rid_map = {r.attrib['Id']: r.attrib['Target'] for r in rels}
        prs = etree.fromstring(z.read('ppt/presentation.xml'))
        ns = {'p': P, 'r': R}
        return ['ppt/' + rid_map[s.attrib[f'{{{R}}}id']]
                for s in prs.find('.//p:sldIdLst', ns)]


def _find_slide(slides_order, files_dict, marker_name):
    """
    Encuentra el primer slide que contiene un shape con el nombre exacto.
    Retorna (index, path). Más robusto que índices fijos: no se rompe
    cuando se duplican slides de perfiles antes del slide de FDA u otros.
    """
    for i, path in enumerate(slides_order):
        root = etree.fromstring(files_dict[path])
        for sp in root.iter(f'{{{P}}}sp'):
            nvpr = sp.find(f'.//{{{P}}}cNvPr')
            if nvpr is not None and nvpr.attrib.get('name') == marker_name:
                return i, path
    raise ValueError(
        f"No se encontró ningún slide con el shape '{marker_name}'. "
        f"Verifica que la plantilla PPTX tenga ese nombre de shape exacto."
    )


# ═══════════════════════════════ manipulación de shapes ═════════════════════

def _hide_shape(sp):
    """
    Hace el shape transparente y vacía su texto.
    Usado en el slide FDA donde los shapes vacíos deben ocultar su relleno
    pero mantener su posición en el layout.
    """
    spPr = sp.find(f'{{{P}}}spPr')
    if spPr is None:
        spPr = etree.SubElement(sp, f'{{{P}}}spPr')

    for fill_tag in ('solidFill', 'gradFill', 'pattFill', 'blipFill', 'grpFill', 'noFill'):
        el = spPr.find(f'{{{A}}}{fill_tag}')
        if el is not None:
            spPr.remove(el)
    etree.SubElement(spPr, f'{{{A}}}noFill')

    ln = spPr.find(f'{{{A}}}ln')
    if ln is None:
        ln = etree.SubElement(spPr, f'{{{A}}}ln')
    for fill_tag in ('solidFill', 'gradFill', 'pattFill', 'noFill'):
        el = ln.find(f'{{{A}}}{fill_tag}')
        if el is not None:
            ln.remove(el)
    etree.SubElement(ln, f'{{{A}}}noFill')

    txb = sp.find(f'{{{P}}}txBody')
    if txb is not None:
        for t in txb.findall(f'.//{{{A}}}t'):
            t.text = ''


def _remove_shape(sp):
    """
    Elimina el shape del árbol XML de forma definitiva.
    Usado para slots de perfiles sin contenido: garantiza que no queden
    tarjetas invisibles, iconos huérfanos ni texto residual en el slide.
    """
    parent = sp.getparent()
    if parent is not None:
        parent.remove(sp)


def _truncate_desc(text, max_chars=DESC_MAX_CHARS):
    """
    Trunca la descripción preservando oraciones completas.
    Estrategia:
      1. Si el texto cabe entero → devolver tal cual.
      2. Buscar el último punto '.' dentro de la ventana max_chars.
         Si existe y está en la primera mitad del texto o más → cortar ahí
         (incluye el punto, sin elipsis: la oración queda completa).
      3. Si no hay punto válido → cortar en el último espacio (límite
         de palabra) y añadir '…' para indicar que hay más.
    No corta en el primer punto; se asegura de conservar el máximo de
    contenido significativo posible.
    """
    if not text or len(text) <= max_chars:
        return text

    window = text[:max_chars + 1]          # +1 para incluir punto exactamente en max_chars
    last_period = window.rfind('.')

    # Usar límite de oración si el punto está en posición >= 20 (para no cortar
    # una oración trivialmente corta al principio del texto)
    if last_period >= 20:
        return text[:last_period + 1]      # incluir el punto; sin elipsis

    # Sin punto útil → cortar en límite de palabra (max_chars-1 para dar espacio al …)
    truncated = text[:max_chars - 1].rsplit(' ', 1)[0]
    return truncated.rstrip('.,;:') + '\u2026'


def _normalize_bodyPr(txb):
    """
    Normaliza el bodyPr del txBody:
    - wrap='square'   → word wrap habilitado
    - normAutofit     → la fuente se reduce ligeramente si el texto no cabe
                        (no expande la caja, evita desbordamiento visual)
    """
    bodyPr = txb.find(f'{{{A}}}bodyPr')
    if bodyPr is None:
        return
    bodyPr.set('wrap', 'square')
    for tag in ('spAutoFit', 'noAutofit', 'normAutofit'):
        el = bodyPr.find(f'{{{A}}}{tag}')
        if el is not None:
            bodyPr.remove(el)
    etree.SubElement(bodyPr, f'{{{A}}}normAutofit')


# ── Constantes para pre-calcular alto de cuadros de descripción ──────────────
# Calibri 12pt en una caja de ~38.5 mm de ancho ≈ 17 chars/línea
_DESC_CHARS_PER_LINE = 17
# 12pt × 110% line spacing × 12700 EMU/pt ≈ 167 640 EMU por línea
_DESC_LINE_HEIGHT_EMU = 167_640
# Margen inferior de seguridad
_DESC_PADDING_EMU = 200_000


def _update_desc_height(sp, desc_text):
    """
    Ajusta el alto (ext.cy) del cuadro de descripción según el contenido.
    Combinado con spAutoFit, garantiza que el texto sea visible
    incluso antes de que PowerPoint haga el primer auto-fit.
    """
    spPr = sp.find(f'{{{P}}}spPr')
    if spPr is None:
        return
    xfrm = spPr.find(f'{{{A}}}xfrm')
    if xfrm is None:
        return
    ext = xfrm.find(f'{{{A}}}ext')
    if ext is None:
        return

    # Calcular líneas necesarias considerando posibles saltos de línea en la desc
    text_lines = [l for l in desc_text.split('\n') if l.strip()] or [desc_text]
    total_lines = sum(
        max(1, -(-len(line) // _DESC_CHARS_PER_LINE))
        for line in text_lines
    )
    needed_cy = total_lines * _DESC_LINE_HEIGHT_EMU + _DESC_PADDING_EMU
    current_cy = int(ext.attrib.get('cy', 0))
    ext.attrib['cy'] = str(max(current_cy, needed_cy))


# ══════════════════════════ posicionamiento de grupos ════════════════════════

def _get_group_off_x(grpSp):
    """Retorna el off.x del grupo externo (grpSpPr/a:xfrm/a:off.x)."""
    grpSpPr = grpSp.find(f'{{{P}}}grpSpPr')
    if grpSpPr is None:
        return None
    xfrm = grpSpPr.find(f'{{{A}}}xfrm')
    if xfrm is None:
        return None
    off = xfrm.find(f'{{{A}}}off')
    if off is None:
        return None
    return int(off.attrib.get('x', 0))


def _set_group_off_x(grpSp, new_x):
    """Actualiza el off.x del grupo externo en el XML."""
    grpSpPr = grpSp.find(f'{{{P}}}grpSpPr')
    if grpSpPr is None:
        return
    xfrm = grpSpPr.find(f'{{{A}}}xfrm')
    if xfrm is None:
        return
    off = xfrm.find(f'{{{A}}}off')
    if off is not None:
        off.attrib['x'] = str(new_x)


def _shift_pic_x(pic, delta):
    """Desplaza en X un elemento pic (avatar) en delta EMU."""
    spPr = pic.find(f'{{{P}}}spPr')
    if spPr is None:
        return
    xfrm = spPr.find(f'{{{A}}}xfrm')
    if xfrm is None:
        return
    off = xfrm.find(f'{{{A}}}off')
    if off is not None:
        off.attrib['x'] = str(int(off.attrib.get('x', 0)) + delta)


def _find_avatar_pics(spTree, groups):
    """
    Asocia cada slot con su pic de avatar al nivel del spTree.
    La asociación se hace por posición X: el centro del pic cae dentro
    del rango horizontal del grupo [off.x, off.x + CARD_WIDTH].
    Retorna un dict {slot_index: (pic_element, original_x)}.
    """
    slot_pics = {}
    for child in list(spTree):
        if child.tag != f'{{{P}}}pic':
            continue
        spPr = child.find(f'{{{P}}}spPr')
        if spPr is None:
            continue
        xfrm = spPr.find(f'{{{A}}}xfrm')
        if xfrm is None:
            continue
        off = xfrm.find(f'{{{A}}}off')
        ext = xfrm.find(f'{{{A}}}ext')
        if off is None:
            continue
        pic_x  = int(off.attrib.get('x', 0))
        pic_cx = int(ext.attrib.get('cx', 0)) if ext is not None else 0
        pic_center = pic_x + pic_cx // 2

        for slot_idx, (grpSp, group_off_x) in enumerate(groups):
            if group_off_x <= pic_center <= group_off_x + _PERFIL_CARD_WIDTH:
                slot_pics[slot_idx] = (child, pic_x)
                break
    return slot_pics


def _find_profile_groups(root):
    """
    Localiza los 4 grpSp externos (tarjetas de perfil) como hijos directos
    del spTree, ordenados por posición X.

    Identificación: un grpSp es una tarjeta de perfil si contiene algún sp
    cuyo nombre coincida con los nombres de rol definidos en _PERFIL_ROLE_NAMES.
    """
    spTree = root.find(f'.//{{{P}}}spTree')
    groups = []
    for child in list(spTree):
        if child.tag != f'{{{P}}}grpSp':
            continue
        inner_names = {
            nvpr.attrib.get('name', '')
            for sp in child.iter(f'{{{P}}}sp')
            for nvpr in [sp.find(f'.//{{{P}}}cNvPr')]
            if nvpr is not None
        }
        if inner_names & _PERFIL_ROLE_NAMES:
            off_x = _get_group_off_x(child)
            groups.append((child, off_x if off_x is not None else 0))
    groups.sort(key=lambda g: g[1])
    return groups


# ═══════════════════════════════ slide perfiles ══════════════════════════════

def _build_para_from_template(template_para, text):
    """Clona el formato de un párrafo plantilla con un nuevo texto.
    El run se inserta ANTES de a:endParaRPr para respetar el orden OOXML."""
    new_para = copy.deepcopy(template_para)
    for r in new_para.findall(f'{{{A}}}r'):
        new_para.remove(r)
    for br in new_para.findall(f'{{{A}}}br'):
        new_para.remove(br)
    rPr = None
    orig_r = template_para.find(f'{{{A}}}r')
    if orig_r is not None:
        orig_rPr = orig_r.find(f'{{{A}}}rPr')
        if orig_rPr is not None:
            rPr = copy.deepcopy(orig_rPr)
    r_elem = etree.Element(f'{{{A}}}r')
    if rPr is not None:
        r_elem.append(rPr)
    t_elem = etree.SubElement(r_elem, f'{{{A}}}t')
    t_elem.text = text
    # Insertar ANTES de endParaRPr (si existe) para OOXML válido
    end_rpr = new_para.find(f'{{{A}}}endParaRPr')
    if end_rpr is not None:
        end_rpr.addprevious(r_elem)
    else:
        new_para.append(r_elem)
    return new_para


def _truncate_to_sentences(text, max_sentences=2):
    """Corta el texto después de max_sentences oraciones (puntos finales)."""
    if not text:
        return text
    count = 0
    for i, c in enumerate(text):
        if c == '.':
            count += 1
            if count >= max_sentences:
                return text[:i + 1].strip()
    return text.strip()


def _edit_perfiles_slide(xml_bytes, perfiles):
    """
    Edita un slide de Perfiles con hasta 4 perfiles.

    Cada slot es un grpSp externo en el spTree. Se opera directamente sobre
    el grupo externo (no sobre shapes individuales) para:
    - Rellenar nombre de rol y descripción dentro del grupo.
    - Eliminar grupos de slots vacíos del spTree por completo
      (spTree.remove(grpSp) en vez de individual _remove_shape, que dejaba
      los shapes locales y rompía el layout del último slide).
    - Recentrar los grupos visibles ajustando su grpSpPr/a:xfrm/a:off.x.
    """
    root     = etree.fromstring(xml_bytes)
    perfiles = perfiles[:4]
    spTree   = root.find(f'.//{{{P}}}spTree')

    groups    = _find_profile_groups(root)          # [(grpSp, off_x), ...] por x
    slot_pics = _find_avatar_pics(spTree, groups)   # {slot_idx: (pic, orig_x)}

    for i, (grpSp, _) in enumerate(groups):
        if i < len(perfiles):
            p = perfiles[i]
            role_name, desc_name = _PERFIL_SLOT_NAMES[i]

            # Mapa de nombres dentro del grupo
            inner_name_map = {}
            for sp in grpSp.iter(f'{{{P}}}sp'):
                nvpr = sp.find(f'.//{{{P}}}cNvPr')
                if nvpr is not None:
                    n = nvpr.attrib.get('name', '')
                    if n not in inner_name_map:
                        inner_name_map[n] = sp

            # Escribir nombre del rol
            sp_nombre = inner_name_map.get(role_name)
            if sp_nombre is not None:
                txb = sp_nombre.find(f'{{{P}}}txBody')
                if txb is not None:
                    all_t = txb.findall(f'.//{{{A}}}t')
                    if all_t:
                        all_t[0].text = p['rol']
                        for t in all_t[1:]:
                            t.text = ''

            # Escribir descripción (máx 2 oraciones para no desbordar el cuadro)
            sp_desc = inner_name_map.get(desc_name)
            if sp_desc is not None:
                txb = sp_desc.find(f'{{{P}}}txBody')
                if txb is not None:
                    desc_bold = p.get('desc_bold', False)
                    raw_desc  = (p['desc'] or '').strip()
                    # No truncar el placeholder "no encontrado"
                    desc_text = raw_desc if desc_bold else _truncate_to_sentences(raw_desc)
                    lines = [l for l in desc_text.split('\n') if l.strip()] or ['']
                    paras = txb.findall(f'{{{A}}}p')
                    template_para = paras[0] if paras else None
                    for para in paras:
                        txb.remove(para)
                    for line in lines:
                        if template_para is not None:
                            new_para = _build_para_from_template(template_para, line)
                            if desc_bold:
                                for r in new_para.findall(f'{{{A}}}r'):
                                    rPr = r.find(f'{{{A}}}rPr')
                                    if rPr is None:
                                        rPr = etree.Element(f'{{{A}}}rPr')
                                        r.insert(0, rPr)
                                    rPr.set('b', '1')
                            txb.append(new_para)
                        else:
                            bold_attr = ' b="1"' if desc_bold else ''
                            p_xml = (f'<a:p xmlns:a="{A}"><a:r>'
                                     f'<a:rPr{bold_attr}/>'
                                     f'<a:t>{_esc(line)}</a:t></a:r></a:p>')
                            txb.append(etree.fromstring(p_xml))
                    _normalize_bodyPr(txb)
        else:
            # Eliminar el grupo completo del spTree (evita tarjetas fantasma)
            spTree.remove(grpSp)
            # Eliminar también el avatar pic de este slot (evita iconos huérfanos)
            slot_pic = slot_pics.get(i)
            if slot_pic is not None:
                pic_el, _ = slot_pic
                spTree.remove(pic_el)

    # Recentrar los grupos visibles y sus avatars si hay menos de 4
    n = len(perfiles)
    if 0 < n < 4:
        total_width = (n - 1) * _PERFIL_PITCH + _PERFIL_CARD_WIDTH
        start_x     = (SLIDE_WIDTH_EMU - total_width) // 2
        for i, (grpSp, old_off_x) in enumerate(groups[:n]):
            new_off_x = start_x + i * _PERFIL_PITCH
            delta = new_off_x - old_off_x
            if delta == 0:
                continue
            _set_group_off_x(grpSp, new_off_x)
            # Desplazar también el avatar para que siga centrado sobre la tarjeta
            slot_pic = slot_pics.get(i)
            if slot_pic is not None:
                pic_el, _ = slot_pic
                _shift_pic_x(pic_el, delta)

    return etree.tostring(root, xml_declaration=True, encoding='UTF-8', standalone=True)


# ═════════════════════════════════ slide FDA ═════════════════════════════════

def _fill_qa_card(sp, lines):
    """
    Llena el CuadroTexto 32 con una lista de líneas de texto.
    Preserva el formato (bullets blancos, tamaño, fuente) del template.
    """
    txb = sp.find(f'{{{P}}}txBody')
    if txb is None:
        return
    paras = txb.findall(f'{{{A}}}p')
    template_para = paras[0] if paras else None
    for para in paras:
        txb.remove(para)
    for line in lines:
        if template_para is not None:
            txb.append(_build_para_from_template(template_para, line))
        else:
            p_xml = (f'<a:p xmlns:a="{A}"><a:r><a:t>{_esc(line)}</a:t></a:r></a:p>')
            txb.append(etree.fromstring(p_xml))
    # spAutoFit para que el cuadro se expanda si hay muchos ítems
    bodyPr = txb.find(f'{{{A}}}bodyPr')
    if bodyPr is not None:
        for tag in ('spAutoFit', 'noAutofit', 'normAutofit'):
            el = bodyPr.find(f'{{{A}}}{tag}')
            if el is not None:
                bodyPr.remove(el)
        etree.SubElement(bodyPr, f'{{{A}}}spAutoFit')


def _edit_fda_slide(xml_bytes, torres, fda_db, usar_genericos=True, incluir_qa=True, items_override=None):
    """
    Edita el slide FDA.
    - items_override: lista de ítems a mostrar directamente (para paginación)
    - usar_genericos=True  (pill ON):  cláusula general siempre
    - usar_genericos=False (pill OFF): ítems específicos de cada torre activa

    Panel QA (cuadro verde derecho — SIEMPRE visible):
    - incluir_qa=True:  mostrar ítems FDA específicos de QA del catálogo
    - incluir_qa=False: mostrar "Las pruebas de calidad no hacen parte..."
    """
    root = etree.fromstring(xml_bytes)

    torres_norm = [_norm(t) for t in torres]

    # ── Ítems columna principal FDA ──────────────────────────────────────────
    if items_override is not None:
        items = items_override
    elif usar_genericos:
        items = CLAUSULA_GENERAL[:6]
    elif len(torres) == 1:
        torre_key = torres_norm[0]
        items = fda_db.get(torre_key, [])
        if not items:
            for k in fda_db:
                if torre_key in k or k in torre_key:
                    items = fda_db[k]
                    break
        items = items[:6] if items else CLAUSULA_GENERAL[:6]
    else:
        items = []
        for torre_key in torres_norm:
            tower_items = fda_db.get(torre_key, [])
            if not tower_items:
                for k in fda_db:
                    if torre_key in k or k in torre_key:
                        tower_items = fda_db[k]
                        break
            for it in tower_items:
                if it not in items:
                    items.append(it)
        items = items[:6] if items else CLAUSULA_GENERAL[:6]

    # ── Ítems panel QA ───────────────────────────────────────────────────────
    if incluir_qa:
        qa_items = fda_db.get('QA', [])
        if not qa_items:
            for k in fda_db:
                if 'QA' in k:
                    qa_items = fda_db[k]
                    break
        qa_lines = qa_items[:4] if qa_items else ['Pruebas de calidad incluidas en el alcance del proyecto.']
    else:
        qa_lines = ['Las pruebas de calidad no hacen parte de esta propuesta.']

    _set_bullet_shapes(root, items)

    for sp in root.iter(f'{{{P}}}sp'):
        nvpr = sp.find(f'.//{{{P}}}cNvPr')
        if nvpr is None:
            continue
        name = nvpr.attrib.get('name', '')

        if name == QA_CARD:
            # Siempre visible: con ítems QA o con mensaje de exclusión
            _fill_qa_card(sp, qa_lines)

        elif name == 'Título 1':
            txb = sp.find(f'{{{P}}}txBody')
            if txb is None:
                continue
            txt = ''.join(t.text or '' for t in txb.findall(f'.//{{{A}}}t')).strip()
            if TITULO_TECNICO in txt:
                if len(torres) == 1:
                    for t in txb.findall(f'.//{{{A}}}t'):
                        if t.text and TITULO_TECNICO in t.text:
                            t.text = f'Fuera del Alcance {torres[0]}'
                else:
                    for t in txb.findall(f'.//{{{A}}}t'):
                        if t.text and TITULO_TECNICO in t.text:
                            t.text = 'Fuera del Alcance General'
            # El título "Pruebas de Calidad" se mantiene siempre visible

    return etree.tostring(root, xml_declaration=True, encoding='UTF-8', standalone=True)


# ══════════════════════════ duplicación de slides ════════════════════════════

def _get_rels_path(slide_path):
    """'ppt/slides/slideN.xml' → 'ppt/slides/_rels/slideN.xml.rels'"""
    parts = slide_path.rsplit('/', 1)
    return f"{parts[0]}/_rels/{parts[1]}.rels"


def _duplicate_perf_slide(files_dict, src_slide_path, insert_after_path):
    """
    Duplica src_slide_path e inserta la copia justo después de insert_after_path
    en el sldIdLst de presentation.xml.
    Retorna el path del nuevo slide.

    IMPORTANTE: esta función modifica presentation.xml para insertar el nuevo
    slide DESPUÉS del slide de referencia, preservando el orden relativo de
    todos los demás slides (FDA, consideraciones, cronograma, etc.).
    Los generadores que buscan slides por contenido (no por índice) no se ven
    afectados por esta inserción.
    """
    NS_REL = 'http://schemas.openxmlformats.org/package/2006/relationships'
    ns_p   = P
    ns_r   = R

    existing_nums = [
        int(re.search(r'slide(\d+)', f).group(1))
        for f in files_dict
        if re.match(r'ppt/slides/slide\d+\.xml$', f)
    ]
    new_num        = max(existing_nums) + 1
    new_slide_path = f'ppt/slides/slide{new_num}.xml'

    # Copiar XML del slide (se sobreescribirá con el contenido correcto justo después)
    files_dict[new_slide_path] = files_dict[src_slide_path]

    # Registrar el nuevo slide en [Content_Types].xml para que PowerPoint no
    # lo rechace como "archivo corrupto". El template usa <Override> por slide.
    CT_NS = 'http://schemas.openxmlformats.org/package/2006/content-types'
    SLIDE_CT = 'application/vnd.openxmlformats-officedocument.presentationml.slide+xml'
    ct_root = etree.fromstring(files_dict['[Content_Types].xml'])
    etree.SubElement(ct_root, f'{{{CT_NS}}}Override', {
        'PartName':    f'/ppt/slides/slide{new_num}.xml',
        'ContentType': SLIDE_CT,
    })
    files_dict['[Content_Types].xml'] = etree.tostring(
        ct_root, xml_declaration=True, encoding='UTF-8', standalone=True
    )

    # Copiar rels del slide (sin notesSlide para evitar conflictos)
    src_rels = _get_rels_path(src_slide_path)
    new_rels = _get_rels_path(new_slide_path)
    if src_rels in files_dict:
        rels_root = etree.fromstring(files_dict[src_rels])
        for rel in rels_root.findall(f'{{{NS_REL}}}Relationship'):
            if 'notesSlide' in rel.attrib.get('Type', ''):
                rels_root.remove(rel)
        files_dict[new_rels] = etree.tostring(
            rels_root, xml_declaration=True, encoding='UTF-8', standalone=True
        )

    # Actualizar presentation.xml.rels con la nueva relación
    prs_rels_path = 'ppt/_rels/presentation.xml.rels'
    prs_rels_root = etree.fromstring(files_dict[prs_rels_path])

    # IMPORTANTE: buscar ref_rid ANTES de añadir la nueva relación,
    # para que la búsqueda no confunda el nuevo slide con el de referencia.
    ref_target = insert_after_path.replace('ppt/', '')
    ref_rid = None
    for rel in prs_rels_root.findall(f'{{{NS_REL}}}Relationship'):
        if rel.attrib.get('Target', '') == ref_target:
            ref_rid = rel.attrib['Id']
            break

    rid_nums = [
        int(m.group(1))
        for r in prs_rels_root.findall(f'{{{NS_REL}}}Relationship')
        for m in [re.search(r'rId(\d+)', r.attrib.get('Id', ''))]
        if m
    ]
    new_rid = f'rId{max(rid_nums) + 1}'
    etree.SubElement(prs_rels_root, f'{{{NS_REL}}}Relationship', {
        'Id':     new_rid,
        'Type':   f'{ns_r}/slide',
        'Target': f'slides/slide{new_num}.xml',
    })
    files_dict[prs_rels_path] = etree.tostring(
        prs_rels_root, xml_declaration=True, encoding='UTF-8', standalone=True
    )

    # Actualizar presentation.xml: insertar sldId justo después del slide de referencia
    prs_path = 'ppt/presentation.xml'
    prs_root = etree.fromstring(files_dict[prs_path])
    sldIdLst = prs_root.find(f'.//{{{ns_p}}}sldIdLst')

    max_sld_id = max(int(s.attrib['id']) for s in sldIdLst)
    new_sld_el = etree.Element(f'{{{ns_p}}}sldId')
    new_sld_el.attrib['id'] = str(max_sld_id + 1)
    new_sld_el.attrib[f'{{{ns_r}}}id'] = new_rid

    children = list(sldIdLst)
    for child in children:
        sldIdLst.remove(child)
    inserted = False
    for child in children:
        sldIdLst.append(child)
        if child.attrib.get(f'{{{ns_r}}}id') == ref_rid and not inserted:
            sldIdLst.append(new_sld_el)
            inserted = True
    if not inserted:
        sldIdLst.append(new_sld_el)

    files_dict[prs_path] = etree.tostring(
        prs_root, xml_declaration=True, encoding='UTF-8', standalone=True
    )

    return new_slide_path


# ═══════════════════════════════ entry point ═════════════════════════════════

def edit(pptx_bytes, config):
    """
    Punto de entrada principal.
    config = {
        filial: str,
        excel_data: { torres: [...], perfiles: [...] },
        torres_seleccionadas: [...],
        opciones: { perfiles: 'excel'|'genericos', fda: 'excel'|'genericos' },
    }
    """
    fda_db, perf_db = _load_generales()

    excel_data     = config.get('excel_data') or {}
    excel_torres   = excel_data.get('torres', [])
    excel_perfiles = excel_data.get('perfiles', [])
    torres_sel     = config.get('torres_seleccionadas', [])
    opciones       = config.get('opciones', {})

    torres_activas = [t['nombre'] for t in excel_torres] if excel_torres else torres_sel

    # ── Determinar perfiles ──────────────────────────────────────────────────
    # El JS envía opciones.perfiles como boolean:
    #   true  → pill "Con genéricos" activada → usar genéricos del DB
    #   false → pill desactivada → usar datos del Excel (Anexos) si existen,
    #            o caer a genéricos si no hay datos en Anexos
    usar_genericos = bool(opciones.get('perfiles'))

    # perfiles_manuales: elegidos desde el buscador del frontend (excelVacio)
    perfiles_manuales = config.get('perfiles_manuales') or []

    if perfiles_manuales:
        base_perfiles = [
            {'rol': p['rol'], 'desc': (p.get('desc') or '').strip()}
            for p in perfiles_manuales
            if p.get('rol')
        ]
        perfiles = _complement_perfiles(base_perfiles, torres_activas, perf_db) if usar_genericos else base_perfiles
    elif excel_perfiles:
        # Buscar descripción en catálogo; si no hay match → placeholder bold
        base_perfiles = []
        for p in excel_perfiles:
            if not p.get('perfil'):
                continue
            catalog_desc = _find_desc_in_catalog(p['perfil'], perf_db)
            if catalog_desc:
                base_perfiles.append({'rol': p['perfil'], 'desc': catalog_desc})
            else:
                base_perfiles.append({'rol': p['perfil'], 'desc': _NO_CATALOG_DESC, 'desc_bold': True})
        perfiles = _complement_perfiles(base_perfiles, torres_activas, perf_db) if usar_genericos else base_perfiles
    else:
        # Sin datos en Anexos ni manual → catálogo completo por torre
        perfiles = []
        for torre in torres_activas:
            key = _norm(torre)
            genericos = perf_db.get(key, [])
            if not genericos:
                for k in perf_db:
                    if key in k or k in key:
                        genericos = perf_db[k]
                        break
            perfiles.extend(genericos)

    # ── Validación y log de paginación ──────────────────────────────────────
    total_perfiles    = len(perfiles)
    slides_necesarios = max(1, -(-total_perfiles // 4))  # ceil division
    chunks            = [perfiles[i:i+4] for i in range(0, max(total_perfiles, 1), 4)]
    print(f'[PERFILES] Torres activas  : {torres_activas}')
    print(f'[PERFILES] Total perfiles  : {total_perfiles}')
    print(f'[PERFILES] Slides necesarios: {slides_necesarios}')
    layout_label = {1: 'centrado único', 2: 'centrado 2 cols',
                    3: 'centrado 3 cols', 4: '4 cols (layout base)'}
    for idx, chunk in enumerate(chunks):
        n = len(chunk)
        print(f'[PERFILES]   Slide {idx + 1}: {n} perfiles → {layout_label.get(n, str(n))}')

    # ── Cargar todos los archivos del PPTX ──────────────────────────────────
    files_dict = {}
    with zipfile.ZipFile(io.BytesIO(pptx_bytes)) as zin:
        files_dict = {name: zin.read(name) for name in zin.namelist()}

    # ── Localizar slides por contenido (no por índice fijo) ─────────────────
    # Esto garantiza que la duplicación de slides de perfiles no desplace
    # el índice del slide FDA ni de otros slides editados por otros generadores.
    slides_order = _get_slide_order(pptx_bytes)

    _, fda_slide_key  = _find_slide(slides_order, files_dict, FDA_MARKER)
    _, perf_slide_key = _find_slide(slides_order, files_dict, PERFILES_MARKER)

    # ── Editar slide(s) FDA ─────────────────────────────────────────────────
    # opciones.fda: true = pill ON = usar genéricos, false = pill OFF = específicos por torre
    usar_genericos_fda = bool(opciones.get('fda', True))

    # Derivar incluir_qa:
    # - Si viene explícito en config (tests), usarlo
    # - Si el frontend envió torres_qa, usar cualquier valor truthy
    # - Si no, QA solo si la torre QA está activa
    if 'incluir_qa' in config:
        incluir_qa = bool(config['incluir_qa'])
    else:
        torres_qa_map = config.get('torres_qa', {})
        incluir_qa = (
            any(v for v in torres_qa_map.values()) or
            any(_norm(t) == 'QA' for t in torres_activas)
        )

    # ── Recolectar ítems FDA ────────────────────────────────────────────────
    # Pill ON = "incluir genéricos" para complementar
    # Pill OFF = solo datos del Excel o catálogo por torre
    FDA_SLIDE_CAP = 6  # bullets que caben en una plantilla FDA
    excel_fda_list = [str(it).strip() for it in (excel_data.get('fda') or []) if str(it).strip()]

    if usar_genericos_fda and not excel_fda_list:
        # Pill ON + sin Excel → CLAUSULA_GENERAL pura
        all_fda_items = list(CLAUSULA_GENERAL)
    elif usar_genericos_fda and excel_fda_list:
        # Pill ON + Excel tiene ítems → complementar
        extras = [g for g in CLAUSULA_GENERAL if g not in excel_fda_list]
        if len(excel_fda_list) < FDA_SLIDE_CAP:
            # Pocos ítems: completar el slide con genéricos
            all_fda_items = excel_fda_list + extras[:FDA_SLIDE_CAP - len(excel_fda_list)]
        else:
            # Slide(s) completos: agregar slide extra con genéricos
            all_fda_items = excel_fda_list + extras
    elif excel_fda_list:
        # Pill OFF + Excel → solo ítems del Excel
        all_fda_items = excel_fda_list
    else:
        # Pill OFF + sin Excel → catálogo por torre, máx 6 (slide único)
        torres_norm_fda = [_norm(t) for t in torres_activas]
        all_fda_items = []
        for torre_key in torres_norm_fda:
            tower_items = fda_db.get(torre_key, [])
            if not tower_items:
                for k in fda_db:
                    if torre_key in k or k in torre_key:
                        tower_items = fda_db[k]
                        break
            for it in tower_items:
                if it not in all_fda_items:
                    all_fda_items.append(it)
        all_fda_items = all_fda_items[:FDA_SLIDE_CAP] if all_fda_items else list(CLAUSULA_GENERAL)

    # Paginar: máx FDA_SLIDE_CAP ítems por slide (coincide con bullets del template)
    fda_chunks = _even_chunks(all_fda_items, max_per=FDA_SLIDE_CAP)

    print(f'[FDA] Torres activas      : {torres_activas}')
    print(f'[FDA] Total ítems FDA     : {len(all_fda_items)}')
    print(f'[FDA] Slides necesarios   : {len(fda_chunks)}')

    fda_template_xml = files_dict[fda_slide_key]

    # Primer slide FDA: editar el original
    files_dict[fda_slide_key] = _edit_fda_slide(
        fda_template_xml, torres_activas, fda_db, usar_genericos_fda, incluir_qa,
        items_override=fda_chunks[0]
    )

    # Slides adicionales: duplicar template limpio e insertar después del anterior
    prev_fda_path = fda_slide_key
    for chunk in fda_chunks[1:]:
        new_path = _duplicate_perf_slide(files_dict, fda_slide_key, prev_fda_path)
        files_dict[new_path] = _edit_fda_slide(
            fda_template_xml, torres_activas, fda_db, usar_genericos_fda, incluir_qa,
            items_override=chunk
        )
        prev_fda_path = new_path

    # ── Editar slide(s) Perfiles ────────────────────────────────────────────
    # Guardar el XML original de la plantilla ANTES de cualquier edición
    # para usarlo como base limpia al duplicar.
    perf_template_xml = files_dict[perf_slide_key]

    # Primer slide: editar el slide original con el primer chunk
    files_dict[perf_slide_key] = _edit_perfiles_slide(perf_template_xml, chunks[0])

    # Slides adicionales: duplicar la plantilla original y editar cada uno
    prev_slide_path = perf_slide_key
    for chunk in chunks[1:]:
        new_path = _duplicate_perf_slide(files_dict, perf_slide_key, prev_slide_path)
        # Usar siempre perf_template_xml (limpio) como base, no el slide ya editado
        files_dict[new_path] = _edit_perfiles_slide(perf_template_xml, chunk)
        prev_slide_path = new_path

    # ── Reconstruir ZIP ─────────────────────────────────────────────────────
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data in files_dict.items():
            zout.writestr(name, data)

    return buf.getvalue()
