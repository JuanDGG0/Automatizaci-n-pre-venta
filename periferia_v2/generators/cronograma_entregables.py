"""
generators/cronograma_entregables.py
Responsable: José

Edita slide 7 (Entregables) del PPTX.
El slide tiene 3 columnas con título de torre + lista de entregables.

Cronograma (slide 9) se deja pendiente — requiere diseño adicional.

Lógica Entregables:
  - Toma entregables de Generales_para_todos.xlsx hoja 'Entregables'
  - Agrupa por torres activas
  - Muestra max 3 torres (una por columna del slide)
  - Si el usuario ingresó entregables manuales, usa esos
"""

import io, zipfile, unicodedata, re
from pathlib import Path
from lxml import etree
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / 'data'
GENERALES = DATA_DIR / 'Generales_para_todos.xlsx'

A = 'http://schemas.openxmlformats.org/drawingml/2006/main'
P = 'http://schemas.openxmlformats.org/presentationml/2006/main'
R = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

# Slide 7: 3 columnas de entregables
# Cada columna: Título 1 (nombre torre) + CuadroTexto (lista)
ENTREGABLES_COLS = [
    {'titulo': 'Título 1', 'lista': 'CuadroTexto 4'},    # columna 1
    {'titulo': 'Título 1', 'lista': 'CuadroTexto 13'},   # columna 2
    {'titulo': 'Título 1', 'lista': 'CuadroTexto 17'},   # columna 3
]

def _norm(s):
    s = (s or '').strip().upper()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', s).strip()

def _esc(t):
    return (t or '').replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')

def _load_entregables(torres_activas):
    """Carga entregables del Excel por torres activas."""
    if not GENERALES.exists():
        return []

    wb = load_workbook(GENERALES)
    ws = wb['Entregables']
    torres_norm = [_norm(t) for t in torres_activas]

    entregables_por_torre = {}
    torre_actual = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]).strip():
            torre_actual = _norm(str(row[0]).strip())
        if row[1] and torre_actual:
            aplica = any(torre_actual in t or t in torre_actual for t in torres_norm)
            if aplica:
                entregables_por_torre.setdefault(torre_actual, []).append(str(row[1]).strip())

    # Devolver lista de {torre, items} para las torres activas (max 3)
    resultado = []
    for torre in torres_activas[:3]:
        key = _norm(torre)
        items = entregables_por_torre.get(key, [])
        if not items:
            for k in entregables_por_torre:
                if key in k or k in key:
                    items = entregables_por_torre[k]
                    break
        resultado.append({'torre': torre, 'items': items[:7]})

    return resultado

def _get_slide_order(pptx_bytes):
    with zipfile.ZipFile(io.BytesIO(pptx_bytes)) as z:
        rels = etree.fromstring(z.read('ppt/_rels/presentation.xml.rels'))
        rid_map = {r.attrib['Id']: r.attrib['Target'] for r in rels}
        prs = etree.fromstring(z.read('ppt/presentation.xml'))
        ns = {'p': P, 'r': R}
        return ['ppt/' + rid_map[s.attrib[f'{{{R}}}id']]
                for s in prs.find('.//p:sldIdLst', ns)]

def _edit_entregables_slide(xml_bytes, entregables_grupos):
    """
    Edita slide 7 con los entregables.
    entregables_grupos = [{'torre': str, 'items': [str, ...]}, ...]
    """
    root = etree.fromstring(xml_bytes)

    # Encontrar los títulos y listas en orden
    titulos = []
    listas = []

    for sp in root.iter(f'{{{P}}}sp'):
        nvpr = sp.find(f'.//{{{P}}}cNvPr')
        if nvpr is None:
            continue
        name = nvpr.attrib.get('name', '')
        txb = sp.find(f'{{{P}}}txBody')
        if txb is None:
            continue
        txt = ''.join(t.text or '' for t in txb.findall(f'.//{{{A}}}t')).strip()

        if name == 'Título 1' and 'Entregables de' in txt:
            titulos.append(sp)
        elif name in ['CuadroTexto 4', 'CuadroTexto 13', 'CuadroTexto 17']:
            listas.append((name, sp))

    # Ordenar listas por nombre
    orden = {'CuadroTexto 4': 0, 'CuadroTexto 13': 1, 'CuadroTexto 17': 2}
    listas.sort(key=lambda x: orden.get(x[0], 99))

    # Editar cada columna
    for i, grupo in enumerate(entregables_grupos[:3]):
        # Título
        if i < len(titulos):
            txb = titulos[i].find(f'{{{P}}}txBody')
            if txb is not None:
                for t in txb.findall(f'.//{{{A}}}t'):
                    if 'Entregables de' in (t.text or ''):
                        t.text = f'Entregables de {grupo["torre"]}'
                        break

        # Lista de entregables
        if i < len(listas):
            _, sp = listas[i]
            txb = sp.find(f'{{{P}}}txBody')
            if txb is None:
                continue

            paras = txb.findall(f'{{{A}}}p')
            for p in paras:
                txb.remove(p)

            for item in grupo['items']:
                p_xml = (
                    f'<a:p xmlns:a="{A}">'
                    f'<a:r><a:t>{_esc(item)}</a:t></a:r>'
                    f'</a:p>'
                )
                txb.append(etree.fromstring(p_xml))

    return etree.tostring(root, xml_declaration=True, encoding='UTF-8', standalone=True)

def edit(pptx_bytes, config):
    """
    Punto de entrada.
    config = {
        filial: str,
        excel_data: { torres: [...] },
        torres_seleccionadas: [...],
        entregables: [...],  # manuales si los hay
        opciones: { entregables: 'genericos'|'manual' }
    }
    """
    excel_data = config.get('excel_data') or {}
    excel_torres = excel_data.get('torres', [])
    torres_sel = config.get('torres_seleccionadas', [])
    opciones = config.get('opciones', {})

    torres_activas = [t['nombre'] for t in excel_torres] if excel_torres else torres_sel

    # Determinar entregables
    if opciones.get('entregables') == 'manual':
        items_manuales = config.get('entregables', [])
        entregables_grupos = [{'torre': 'Proyecto', 'items': items_manuales[:7]}]
    else:
        entregables_grupos = _load_entregables(torres_activas)

    # Editar slides
    slides_order = _get_slide_order(pptx_bytes)

    files_dict = {}
    with zipfile.ZipFile(io.BytesIO(pptx_bytes)) as zin:
        files_dict = {name: zin.read(name) for name in zin.namelist()}

    # Slide 7 → Entregables (índice 6)
    ent_slide_key = slides_order[6]
    files_dict[ent_slide_key] = _edit_entregables_slide(
        files_dict[ent_slide_key],
        entregables_grupos
    )

    # Slide 9 → Cronograma (índice 8) — TODO: pendiente implementar
    # Por ahora se deja con el contenido del template

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data in files_dict.items():
            zout.writestr(name, data)

    return buf.getvalue()
