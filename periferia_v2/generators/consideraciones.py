"""
generators/consideraciones.py
Responsable: Juan

Edita el slide de Consideraciones del PPTX de la filial.

Estructura del slide:
  - Cada consideración es un grpSp (grupo) que contiene:
      · El shape redondeado (Redondear rectángulo de esquina diagonal 14)
      · El ícono ⓘ (pic)
  - El slide template tiene 4 grupos (Grupo 2, Grupo 9, Grupo 12, Grupo 15)
  - El color del template es un gradiente bg1 → E4EFD9: NO se toca

Lógica:
  - Las consideraciones del Excel de estimación (columna J) se incluyen SIEMPRE.
  - Pill ON: además se agregan los genéricos de Generales_para_todos.xlsx
             filtrados por torre, al final y sin duplicar.
  - Pill OFF: solo las del Excel de estimación.
  - Máximo 5 consideraciones por slide.
  - Si el texto supera 3 líneas (~159 chars), el grupo se alarga lo necesario.
  - Si no supera el límite, el tamaño del grupo NO se toca.
  - Los grupos se posicionan secuencialmente: cada uno empieza donde termina el anterior.
  - Grupos sobrantes se eliminan del XML por completo.
  - El nombre de la filial se renderiza en negrita dentro del texto.
  - Si hay más de 5 consideraciones, se duplica el slide.
  - Slide localizado por contenido (>= 4 grpSp con SHAPE_NAME), no por índice.
"""

import copy, io, math, re, unicodedata, zipfile
from pathlib import Path
from lxml import etree
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / 'data'
GENERALES = DATA_DIR / 'Generales_para_todos.xlsx'

A = 'http://schemas.openxmlformats.org/drawingml/2006/main'
P = 'http://schemas.openxmlformats.org/presentationml/2006/main'
R = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

SHAPE_NAME = 'Redondear rectángulo de esquina diagonal 14'

PLACEHOLDER_CLIENTE = 'XXXXXXXXXX'
PLACEHOLDER_FILIAL  = 'Filial'
MAX_POR_SLIDE       = 5

# Constantes de layout (medidas del template)
Y_START        = 850000   # Y del primer grupo, debajo del título
GRUPO_CY_BASE  = 708641   # altura base del grupo externo (EMU)
SHAPE_CY_BASE  = 831267   # altura base del shape interno (EMU)
GAP_ORIGINAL   = 162777   # gap entre grupos en el template

# Calibración de texto: el placeholder de ~160 chars cabe en 3 líneas en el shape base
CHARS_POR_LINEA = 53      # chars por línea a 12pt Calibri italic en ese ancho
LINEAS_BASE     = 3       # líneas que caben en el shape/grupo sin agrandar
EMU_POR_LINEA   = GRUPO_CY_BASE // LINEAS_BASE  # 236213 EMU por línea

FILIAL_NOMBRES = {
    'corp':  'Periferia IT Corp',
    'group': 'Periferia IT Group',
    'cbit':  'Contact & Business IT',
}


# ═══════════════════════════════ utilidades ══════════════════════════════════

def _norm(s):
    s = (s or '').strip().upper()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', s).strip()


def _apply_replacements(texto, cliente, filial_nombre):
    texto = texto.replace(PLACEHOLDER_CLIENTE, cliente or 'el cliente')
    texto = texto.replace(PLACEHOLDER_FILIAL,  filial_nombre)
    return texto


def _calc_delta(texto):
    """
    Calcula cuánto hay que alargar el grupo (en EMU) para que el texto quepa.
    Retorna 0 si el texto cabe en las LINEAS_BASE sin agrandar.
    """
    lineas = math.ceil(len(texto) / CHARS_POR_LINEA)
    if lineas <= LINEAS_BASE:
        return 0
    return (lineas - LINEAS_BASE) * EMU_POR_LINEA


# ═══════════════════════════════ fuentes de datos ════════════════════════════

def _load_desde_excel(excel_consideraciones, cliente, filial_nombre):
    """
    Carga las consideraciones del Excel de estimación (columna J).
    Se incluyen SIEMPRE, independientemente de la pill.
    """
    if not excel_consideraciones:
        print('[CONSIDERACIONES] excel_data.consideraciones vacío.')
        return []

    resultado = []
    for item in excel_consideraciones:
        if not item or not str(item).strip():
            continue
        texto = _apply_replacements(str(item).strip(), cliente, filial_nombre)
        if texto not in resultado:
            resultado.append(texto)
    return resultado


def _load_desde_generales(torres_activas, cliente, filial_nombre):
    """
    Pill ON: genéricos de Generales_para_todos.xlsx filtrados por torre.
    Se agregan al final de las del Excel, sin duplicar.
    """
    if not GENERALES.exists():
        print('[CONSIDERACIONES] Advertencia: Generales_para_todos.xlsx no encontrado.')
        return []

    wb           = load_workbook(GENERALES)
    ws           = wb['Consideraciones']
    torres_norm  = {_norm(t) for t in torres_activas}
    torre_actual = None
    resultado    = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        col_a = row[0]
        col_b = row[1] if len(row) > 1 else None

        if col_a and str(col_a).strip():
            torre_actual = _norm(str(col_a).strip().replace('TORRE ', ''))

        if not (col_b and str(col_b).strip() and torre_actual):
            continue

        aplica = any(torre_actual in t or t in torre_actual for t in torres_norm)
        if not aplica:
            continue

        texto = _apply_replacements(str(col_b).strip(), cliente, filial_nombre)
        if texto not in resultado:
            resultado.append(texto)

    return resultado


# ═══════════════════════════════ manipulación XML ════════════════════════════

def _find_grupos(root):
    """
    Retorna lista de grpSp que contienen un shape con SHAPE_NAME,
    ordenados por posición Y (orden visual de arriba a abajo).
    """
    spTree = root.find(f'.//{{{P}}}spTree')
    grupos = []
    for child in list(spTree):
        if child.tag != f'{{{P}}}grpSp':
            continue
        tiene_shape = any(
            nvpr.attrib.get('name', '') == SHAPE_NAME
            for sp   in child.iter(f'{{{P}}}sp')
            for nvpr in [sp.find(f'.//{{{P}}}cNvPr')]
            if nvpr is not None
        )
        if tiene_shape:
            grpSpPr = child.find(f'{{{P}}}grpSpPr')
            xfrm    = grpSpPr.find(f'{{{A}}}xfrm') if grpSpPr is not None else None
            off     = xfrm.find(f'{{{A}}}off') if xfrm is not None else None
            y       = int(off.attrib.get('y', 0)) if off is not None else 0
            grupos.append((y, child))

    grupos.sort(key=lambda g: g[0])
    return [g for _, g in grupos]


def _write_text_in_grupo(grpSp, texto, filial_nombre=''):
    """
    Escribe texto en el shape redondeado dentro del grupo.
    - Pone en negrita el nombre de la filial si aparece en el texto.
    - Si el texto supera LINEAS_BASE líneas, alarga el grupo y el shape.
    - Si no supera el límite, NO modifica las dimensiones.
    Retorna el delta de altura aplicado al grupo (0 si no cambió).
    """
    delta = _calc_delta(texto)

    for sp in grpSp.iter(f'{{{P}}}sp'):
        nvpr = sp.find(f'.//{{{P}}}cNvPr')
        if nvpr is None or nvpr.attrib.get('name', '') != SHAPE_NAME:
            continue
        txb = sp.find(f'{{{P}}}txBody')
        if txb is None:
            continue

        paras = txb.findall(f'{{{A}}}p')
        if not paras:
            break
        para = paras[0]

        # Capturar rPr base del primer run existente
        base_rPr = None
        orig_r = para.find(f'{{{A}}}r')
        if orig_r is not None:
            orig_rPr = orig_r.find(f'{{{A}}}rPr')
            if orig_rPr is not None:
                base_rPr = copy.deepcopy(orig_rPr)

        # Limpiar runs y saltos existentes
        for r in para.findall(f'{{{A}}}r'):
            para.remove(r)
        for br in para.findall(f'{{{A}}}br'):
            para.remove(br)

        def _make_run(txt, bold=False):
            r_el = etree.SubElement(para, f'{{{A}}}r')
            rPr  = copy.deepcopy(base_rPr) if base_rPr is not None else etree.Element(f'{{{A}}}rPr')
            if bold:
                rPr.set('b', '1')
            else:
                rPr.attrib.pop('b', None)
            r_el.insert(0, rPr)
            t_el = etree.SubElement(r_el, f'{{{A}}}t')
            t_el.text = txt
            if txt.startswith(' ') or txt.endswith(' '):
                t_el.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')

        # Negrita en el nombre de la filial
        if filial_nombre and filial_nombre in texto:
            idx    = texto.index(filial_nombre)
            before = texto[:idx]
            after  = texto[idx + len(filial_nombre):]
            if before:
                _make_run(before, bold=False)
            _make_run(filial_nombre, bold=True)
            if after:
                _make_run(after, bold=False)
        else:
            _make_run(texto, bold=False)

        # Alargar shape y grupo SOLO si el texto no cabe en el tamaño base
        if delta > 0:
            # Alargar shape interno
            spPr = sp.find(f'{{{P}}}spPr')
            if spPr is not None:
                xfrm = spPr.find(f'{{{A}}}xfrm')
                if xfrm is not None:
                    ext = xfrm.find(f'{{{A}}}ext')
                    if ext is not None:
                        ext.attrib['cy'] = str(SHAPE_CY_BASE + delta)

            # Alargar grupo externo (ext.cy y chExt.cy)
            grpSpPr = grpSp.find(f'{{{P}}}grpSpPr')
            if grpSpPr is not None:
                xfrm_grp = grpSpPr.find(f'{{{A}}}xfrm')
                if xfrm_grp is not None:
                    ext_grp = xfrm_grp.find(f'{{{A}}}ext')
                    if ext_grp is not None:
                        ext_grp.attrib['cy'] = str(GRUPO_CY_BASE + delta)
                    chext = xfrm_grp.find(f'{{{A}}}chExt')
                    if chext is not None:
                        chext.attrib['cy'] = str(SHAPE_CY_BASE + delta)

        return delta

    return 0


def _remove_grupo(root, grpSp):
    """Elimina el grpSp completo del spTree."""
    spTree = root.find(f'.//{{{P}}}spTree')
    if spTree is not None:
        spTree.remove(grpSp)


def _set_grupo_y(grpSp, new_y):
    """Actualiza solo la posición Y externa del grupo en grpSpPr."""
    grpSpPr = grpSp.find(f'{{{P}}}grpSpPr')
    xfrm    = grpSpPr.find(f'{{{A}}}xfrm') if grpSpPr is not None else None
    off     = xfrm.find(f'{{{A}}}off') if xfrm is not None else None
    if off is not None:
        off.attrib['y'] = str(new_y)


# ═══════════════════════════════ localización y duplicación de slide ══════════

def _get_slide_order(pptx_bytes):
    with zipfile.ZipFile(io.BytesIO(pptx_bytes)) as z:
        rels    = etree.fromstring(z.read('ppt/_rels/presentation.xml.rels'))
        rid_map = {r.attrib['Id']: r.attrib['Target'] for r in rels}
        prs     = etree.fromstring(z.read('ppt/presentation.xml'))
        ns      = {'p': P, 'r': R}
        return ['ppt/' + rid_map[s.attrib[f'{{{R}}}id']]
                for s in prs.find('.//p:sldIdLst', ns)]


def _cuenta_grupos_con_shape(path, files_dict):
    """Cuenta grpSp con SHAPE_NAME en un slide."""
    root   = etree.fromstring(files_dict[path])
    spTree = root.find(f'.//{{{P}}}spTree')
    if spTree is None:
        return 0
    return sum(
        1
        for child in list(spTree)
        if child.tag == f'{{{P}}}grpSp' and any(
            nvpr.attrib.get('name', '') == SHAPE_NAME
            for sp   in child.iter(f'{{{P}}}sp')
            for nvpr in [sp.find(f'.//{{{P}}}cNvPr')]
            if nvpr is not None
        )
    )


def _find_cons_slide(slides_order, files_dict):
    """Localiza el slide con >= 4 grpSp que contengan SHAPE_NAME."""
    for path in slides_order:
        if _cuenta_grupos_con_shape(path, files_dict) >= 4:
            return path

    fallback_idx = min(9, len(slides_order) - 1)
    print(f'[CONSIDERACIONES] Advertencia: slide no encontrado. Usando índice {fallback_idx}.')
    return slides_order[fallback_idx]


def _duplicate_slide(files_dict, src_path, insert_after_path):
    """
    Duplica src_path e inserta la copia justo después de insert_after_path.
    Retorna el path del nuevo slide.
    """
    NS_REL   = 'http://schemas.openxmlformats.org/package/2006/relationships'
    SLIDE_CT = 'application/vnd.openxmlformats-officedocument.presentationml.slide+xml'

    existing_nums = [
        int(re.search(r'slide(\d+)', f).group(1))
        for f in files_dict
        if re.match(r'ppt/slides/slide\d+\.xml$', f)
    ]
    new_num        = max(existing_nums) + 1
    new_slide_path = f'ppt/slides/slide{new_num}.xml'
    files_dict[new_slide_path] = files_dict[src_path]

    CT_NS   = 'http://schemas.openxmlformats.org/package/2006/content-types'
    ct_root = etree.fromstring(files_dict['[Content_Types].xml'])
    etree.SubElement(ct_root, f'{{{CT_NS}}}Override', {
        'PartName':    f'/ppt/slides/slide{new_num}.xml',
        'ContentType': SLIDE_CT,
    })
    files_dict['[Content_Types].xml'] = etree.tostring(
        ct_root, xml_declaration=True, encoding='UTF-8', standalone=True)

    def _rels_path(p):
        parts = p.rsplit('/', 1)
        return f"{parts[0]}/_rels/{parts[1]}.rels"

    src_rels = _rels_path(src_path)
    new_rels = _rels_path(new_slide_path)
    if src_rels in files_dict:
        rels_root = etree.fromstring(files_dict[src_rels])
        for rel in rels_root.findall(f'{{{NS_REL}}}Relationship'):
            if 'notesSlide' in rel.attrib.get('Type', ''):
                rels_root.remove(rel)
        files_dict[new_rels] = etree.tostring(
            rels_root, xml_declaration=True, encoding='UTF-8', standalone=True)

    prs_rels_path = 'ppt/_rels/presentation.xml.rels'
    prs_rels_root = etree.fromstring(files_dict[prs_rels_path])

    ref_target = insert_after_path.replace('ppt/', '')
    ref_rid    = next((
        r.attrib['Id']
        for r in prs_rels_root.findall(f'{{{NS_REL}}}Relationship')
        if r.attrib.get('Target', '') == ref_target
    ), None)

    rid_nums = [
        int(m.group(1))
        for r in prs_rels_root.findall(f'{{{NS_REL}}}Relationship')
        for m in [re.search(r'rId(\d+)', r.attrib.get('Id', ''))]
        if m
    ]
    new_rid = f'rId{max(rid_nums) + 1}'
    etree.SubElement(prs_rels_root, f'{{{NS_REL}}}Relationship', {
        'Id':     new_rid,
        'Type':   f'{R}/slide',
        'Target': f'slides/slide{new_num}.xml',
    })
    files_dict[prs_rels_path] = etree.tostring(
        prs_rels_root, xml_declaration=True, encoding='UTF-8', standalone=True)

    prs_root = etree.fromstring(files_dict['ppt/presentation.xml'])
    sldIdLst = prs_root.find(f'.//{{{P}}}sldIdLst')
    max_id   = max(int(s.attrib['id']) for s in sldIdLst)
    new_sld  = etree.Element(f'{{{P}}}sldId')
    new_sld.attrib['id']         = str(max_id + 1)
    new_sld.attrib[f'{{{R}}}id'] = new_rid

    children = list(sldIdLst)
    for child in children:
        sldIdLst.remove(child)
    inserted = False
    for child in children:
        sldIdLst.append(child)
        if child.attrib.get(f'{{{R}}}id') == ref_rid and not inserted:
            sldIdLst.append(new_sld)
            inserted = True
    if not inserted:
        sldIdLst.append(new_sld)

    files_dict['ppt/presentation.xml'] = etree.tostring(
        prs_root, xml_declaration=True, encoding='UTF-8', standalone=True)

    return new_slide_path


# ═══════════════════════════════ edición del slide ═══════════════════════════

def _edit_cons_slide(xml_bytes, chunk, filial_nombre=''):
    """
    Edita un slide con 1–5 consideraciones.
    - Si hay 5, duplica el último grupo del template.
    - Posiciona grupos secuencialmente: cada uno arranca donde termina el anterior + GAP_ORIGINAL.
    - Los grupos con texto largo se alargan; los siguientes se desplazan automáticamente.
    - Los grupos con texto corto NO se modifican en tamaño.
    - Elimina grupos sobrantes.
    """
    root   = etree.fromstring(xml_bytes)
    grupos = _find_grupos(root)

    n_cons = len(chunk)

    # Si necesitamos 5 grupos, duplicar el último del template
    if n_cons > len(grupos):
        template_grp = grupos[-1]
        new_grp      = copy.deepcopy(template_grp)
        existing_ids = [
            int(el.attrib['id'])
            for el in root.iter()
            if 'id' in el.attrib and el.attrib['id'].isdigit()
        ]
        next_id = max(existing_ids, default=0) + 1
        for el in new_grp.iter():
            if 'id' in el.attrib and el.attrib['id'].isdigit():
                el.attrib['id'] = str(next_id)
                next_id += 1
        spTree = root.find(f'.//{{{P}}}spTree')
        if spTree is not None:
            spTree.append(new_grp)
        grupos.append(new_grp)

    # Posicionamiento secuencial: el Y de cada grupo depende del cy real del anterior
    current_y = Y_START

    for i, grpSp in enumerate(grupos):
        if i < n_cons:
            _set_grupo_y(grpSp, current_y)
            delta = _write_text_in_grupo(grpSp, chunk[i], filial_nombre)
            # El siguiente grupo arranca después de este (cy real) + gap
            current_y += GRUPO_CY_BASE + delta + GAP_ORIGINAL
        else:
            _remove_grupo(root, grpSp)

    return etree.tostring(root, xml_declaration=True, encoding='UTF-8', standalone=True)


# ═══════════════════════════════ entry point ═════════════════════════════════

def edit(pptx_bytes, config):
    """
    config = {
        filial: str,
        excel_data: {
            torres: [{ nombre: str }, ...],
            cliente: str,
            consideraciones: [str, ...],   # columna J parseada por el frontend
        },
        torres_seleccionadas: [...],
        opciones: { consideraciones: bool },  # true = pill ON → agregar genéricos
    }
    """
    excel_data            = config.get('excel_data') or {}
    excel_torres          = excel_data.get('torres', [])
    torres_sel            = config.get('torres_seleccionadas', [])
    filial_key            = config.get('filial', 'corp')
    cliente               = excel_data.get('cliente') or config.get('cliente', '')
    opciones              = config.get('opciones', {})
    usar_genericos        = bool(opciones.get('consideraciones', False))
    excel_consideraciones = excel_data.get('consideraciones', [])

    torres_activas = [t['nombre'] for t in excel_torres] if excel_torres else torres_sel
    filial_nombre  = FILIAL_NOMBRES.get(filial_key, 'Periferia IT')

    print(f'[CONSIDERACIONES] Torres activas  : {torres_activas}')
    print(f'[CONSIDERACIONES] Cliente         : {cliente}')
    print(f'[CONSIDERACIONES] Filial          : {filial_nombre}')
    print(f'[CONSIDERACIONES] Pill genéricos  : {"ON" if usar_genericos else "OFF"}')

    # Las del Excel de estimación se incluyen SIEMPRE
    consideraciones = _load_desde_excel(excel_consideraciones, cliente, filial_nombre)
    print(f'[CONSIDERACIONES] Desde Excel     : {len(consideraciones)}')

    # Pill ON → agregar genéricos al final, sin duplicar
    if usar_genericos:
        genericos = _load_desde_generales(torres_activas, cliente, filial_nombre)
        for item in genericos:
            if item not in consideraciones:
                consideraciones.append(item)
        print(f'[CONSIDERACIONES] + Genéricos     : {len(genericos)}')

    print(f'[CONSIDERACIONES] Total           : {len(consideraciones)}')

    chunks = [consideraciones[i:i+MAX_POR_SLIDE]
              for i in range(0, max(len(consideraciones), 1), MAX_POR_SLIDE)]

    files_dict = {}
    with zipfile.ZipFile(io.BytesIO(pptx_bytes)) as zin:
        files_dict = {name: zin.read(name) for name in zin.namelist()}

    slides_order   = _get_slide_order(pptx_bytes)
    cons_slide_key = _find_cons_slide(slides_order, files_dict)
    template_xml   = files_dict[cons_slide_key]

    files_dict[cons_slide_key] = _edit_cons_slide(template_xml, chunks[0], filial_nombre)

    prev_path = cons_slide_key
    for chunk in chunks[1:]:
        new_path = _duplicate_slide(files_dict, cons_slide_key, prev_path)
        files_dict[new_path] = _edit_cons_slide(template_xml, chunk, filial_nombre)
        prev_path = new_path

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data in files_dict.items():
            zout.writestr(name, data)

    return buf.getvalue()