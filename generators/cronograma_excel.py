"""
cronograma_excel.py
Genera el cronograma como XLSX usando shapes roundRect (drawingML),
replicando el estilo de la plantilla FABRICA_PR-FR-004.

Lógica de layout:
  - Si total_semanas <= 24 (<= 6 meses): muestra Meses + Semanas + Sprints (completo)
  - Si total_semanas > 24 (> 6 meses):  muestra solo Meses + Sprints (sin fila de Semanas)
"""

import io
import math
import zipfile
from lxml import etree
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ─── Constantes ───────────────────────────────────────────────────────────────
HORAS_SEMANALES    = 43   # horas laborales por persona por semana
SEMANAS_POR_MES    = 4
SEMANAS_POR_SPRINT = 2
UMBRAL_MESES       = 24   # semanas; si se supera → mostrar fila de Meses

# Colores de barras por torre (ciclo)
BARRA_COLORES = [
    "0DC56D",  # verde
    "2FF195",  # verde claro
    "24BABA",  # teal
    "2F2BCB",  # azul/morado
    "1D1B80",  # azul oscuro
    "FF6D00",  # naranja
    "D50000",  # rojo
]

# Colores del header
COLOR_MES     = "757070"   # gris oscuro
COLOR_SEMANA  = "D0CECE"   # gris claro
COLOR_SPRINT  = "798EA9"   # azul grisáceo
COLOR_KICKOFF = "757070"

# EMU — todos los offsets se calculan a partir de estos
PADDING        = 38_100      # margen interno de shapes
COL_A_EMU      = 1_781_175   # columna etiqueta torre (26 chars)
COL_SEMANA_EMU =   414_337   # columna de semana (5.5 chars)
ROW_HDR_EMU    =   304_800   # fila de header  (24 pt)
ROW_SUB_EMU    =   254_000   # filas semana/sprint (20 pt)
ROW_ACT_EMU    =   279_400   # filas de actividad (22 pt)


# ─── API pública ──────────────────────────────────────────────────────────────
def generate_cronograma(config: dict) -> bytes:
    actividades = config.get("actividades", [])
    if not actividades:
        raise ValueError("No hay actividades para generar cronograma")

    for act in actividades:
        personas = max(1, int(act.get("personas", 1)))
        act["semanas"] = max(1, math.ceil(act["horas"] / personas / HORAS_SEMANALES))

    total_semanas = max(a["semanas"] for a in actividades)
    sin_semanas   = total_semanas > UMBRAL_MESES   # largo: ocultar fila de semanas

    # Número de filas de header: 3 (mes+semana+sprint) o 2 (mes+sprint, sin semanas)
    n_hdr = 2 if sin_semanas else 3

    wb = Workbook()
    ws = wb.active
    ws.title = "Cronograma"
    ws.sheet_view.showGridLines = False

    _configurar_dimensiones(ws, total_semanas, n_hdr)

    buffer = io.BytesIO()
    wb.save(buffer)

    return _inyectar_drawing(buffer.getvalue(), actividades, total_semanas, sin_semanas, n_hdr)


# ─── Dimensiones ──────────────────────────────────────────────────────────────
def _configurar_dimensiones(ws, total_semanas: int, n_hdr: int):
    # Filas de header
    for r in range(1, n_hdr + 1):
        ws.row_dimensions[r].height = 24 if r == 1 else 20

    # Filas de actividades
    for r in range(n_hdr + 1, n_hdr + 1 + 40):
        ws.row_dimensions[r].height = 22

    # Columnas
    ws.column_dimensions['A'].width = 26                         # etiqueta torre
    ws.column_dimensions['B'].width = 5.5                        # Kick Off
    for i in range(total_semanas):                               # S1..Sn
        ws.column_dimensions[get_column_letter(i + 3)].width = 5.5


# ─── Inyección del drawing ─────────────────────────────────────────────────────
def _inyectar_drawing(xlsx_bytes: bytes, actividades: list,
                      total_semanas: int, sin_semanas: bool, n_hdr: int) -> bytes:
    drawing_xml = _build_drawing_xml(actividades, total_semanas, sin_semanas, n_hdr)

    with zipfile.ZipFile(io.BytesIO(xlsx_bytes), 'r') as zin:
        files = {n: zin.read(n) for n in zin.namelist()}

    files['xl/drawings/drawing1.xml'] = drawing_xml.encode('utf-8')

    sheet_rel_path = 'xl/worksheets/_rels/sheet1.xml.rels'
    if sheet_rel_path not in files:
        files[sheet_rel_path] = (
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            b'<Relationship Id="rId10" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"'
            b' Target="../drawings/drawing1.xml"/>'
            b'</Relationships>'
        )
    else:
        rel_xml = files[sheet_rel_path].decode('utf-8')
        if 'drawing1.xml' not in rel_xml:
            rel_xml = rel_xml.replace(
                '</Relationships>',
                '<Relationship Id="rId10" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"'
                ' Target="../drawings/drawing1.xml"/></Relationships>'
            )
            files[sheet_rel_path] = rel_xml.encode('utf-8')

    sheet_xml = files['xl/worksheets/sheet1.xml'].decode('utf-8')
    if '<drawing ' not in sheet_xml:
        if 'xmlns:r=' not in sheet_xml:
            sheet_xml = sheet_xml.replace(
                '<worksheet ',
                '<worksheet xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" ',
                1
            )
        sheet_xml = sheet_xml.replace('</worksheet>', '<drawing r:id="rId10"/></worksheet>')
        files['xl/worksheets/sheet1.xml'] = sheet_xml.encode('utf-8')

    ct = files['[Content_Types].xml'].decode('utf-8')
    if 'drawing1.xml' not in ct:
        ct = ct.replace(
            '</Types>',
            '<Override PartName="/xl/drawings/drawing1.xml"'
            ' ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/></Types>'
        )
        files['[Content_Types].xml'] = ct.encode('utf-8')

    out = io.BytesIO()
    with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data in files.items():
            zout.writestr(name, data)
    return out.getvalue()


# ─── Construcción del XML de shapes ───────────────────────────────────────────
def _build_drawing_xml(actividades: list, total_semanas: int,
                       sin_semanas: bool, n_hdr: int) -> str:
    XDR = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
    A   = 'http://schemas.openxmlformats.org/drawingml/2006/main'

    root = etree.Element(f'{{{XDR}}}wsDr', nsmap={
        'xdr': XDR, 'a': A,
        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    })

    sid = 1

    # Índices de filas de header (0-based para drawingML):
    #   Normal (<= 6 meses): row 0 = Meses, row 1 = Semanas, row 2 = Sprints
    #   Largo  (> 6 meses):  row 0 = Meses, row 1 = Sprints  (sin semanas)
    ROW_MES    = 0
    if sin_semanas:
        ROW_SEMANA = None   # no se renderiza
        ROW_SPRINT = 1
    else:
        ROW_SEMANA = 1
        ROW_SPRINT = 2

    # col 0 = A (etiquetas torres), col 1 = Kick Off, col 2..N+1 = S1..SN
    COL_KICKOFF = 1
    COL_S1      = 2   # primera semana real

    # ── Kick Off ─────────────────────────────────────────────────────────
    # Abarca desde ROW_MES hasta ROW_SPRINT en su columna
    sid = _shape(root, sid, "Kick Off",
                 COL_KICKOFF, ROW_MES, COL_KICKOFF, ROW_SPRINT,
                 COLOR_KICKOFF, "default", "FFFFFF", 900, True)

    # ── Fila de Meses (siempre presente) ─────────────────────────────────
    col_cur = COL_S1
    for m in range(math.ceil(total_semanas / SEMANAS_POR_MES)):
        col_end = min(col_cur + SEMANAS_POR_MES - 1, COL_S1 + total_semanas - 1)
        sid = _shape(root, sid, f"Mes {m + 1}",
                     col_cur, ROW_MES, col_end, ROW_MES,
                     COLOR_MES, "default", "FFFFFF", 900, True,
                     row_height_emu=ROW_HDR_EMU)
        col_cur += SEMANAS_POR_MES

    # ── Fila de Semanas (solo si <= 6 meses) ─────────────────────────────
    if not sin_semanas:
        for s in range(total_semanas):
            col = COL_S1 + s
            sid = _shape(root, sid, f"S{s + 1}",
                         col, ROW_SEMANA, col, ROW_SEMANA,
                         COLOR_SEMANA, "default", "44546A", 800,
                         row_height_emu=ROW_SUB_EMU)

    # ── Fila de Sprints ───────────────────────────────────────────────────
    sprint_n = 1
    for s in range(0, total_semanas, SEMANAS_POR_SPRINT):
        col_start = COL_S1 + s
        col_end   = min(col_start + SEMANAS_POR_SPRINT - 1, COL_S1 + total_semanas - 1)
        sid = _shape(root, sid, f"Sprint {sprint_n}",
                     col_start, ROW_SPRINT, col_end, ROW_SPRINT,
                     COLOR_SPRINT, "default", "FFFFFF", 800,
                     row_height_emu=ROW_SUB_EMU)
        sprint_n += 1

    # ── Actividades ───────────────────────────────────────────────────────
    for i, act in enumerate(actividades):
        row   = n_hdr + i   # 0-based row index
        color = BARRA_COLORES[i % len(BARRA_COLORES)]

        # Etiqueta torre (columna A = col 0)
        sid = _shape(root, sid, act["torre"],
                     0, row, 0, row,
                     color, "17948", "FFFFFF", 900, True,
                     col_width_emu=COL_A_EMU, row_height_emu=ROW_ACT_EMU)

        # Barra de duración
        bar_end = COL_S1 + act["semanas"] - 1
        sid = _shape(root, sid, "",
                     COL_S1, row, bar_end, row,
                     color, "50000", "FFFFFF", 900,
                     row_height_emu=ROW_ACT_EMU)

    return etree.tostring(root, xml_declaration=True, encoding='UTF-8', standalone=True).decode('utf-8')


# ─── Helper: una shape twoCellAnchor ──────────────────────────────────────────
def _shape(parent, sid: int, text: str,
           col_from: int, row_from: int,
           col_to:   int, row_to:   int,
           color: str,
           adj:        str = "default",
           text_color: str = "FFFFFF",
           font_size:  int = 900,
           bold:       bool = False,
           col_width_emu:  int = COL_SEMANA_EMU,
           row_height_emu: int = ROW_ACT_EMU) -> int:

    XDR = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
    A   = 'http://schemas.openxmlformats.org/drawingml/2006/main'

    anchor = etree.SubElement(parent, f'{{{XDR}}}twoCellAnchor', editAs='oneCell')

    # ── from: borde superior-izquierdo con PADDING interior
    frm = etree.SubElement(anchor, f'{{{XDR}}}from')
    etree.SubElement(frm, f'{{{XDR}}}col').text    = str(col_from)
    etree.SubElement(frm, f'{{{XDR}}}colOff').text = str(PADDING)
    etree.SubElement(frm, f'{{{XDR}}}row').text    = str(row_from)
    etree.SubElement(frm, f'{{{XDR}}}rowOff').text = str(PADDING)

    # ── to: borde inferior-derecho
    # El offset "to" marca hasta dónde llega la shape DENTRO de la celda col_to/row_to.
    # Siempre queremos llegar al borde derecho de col_to y al borde inferior de row_to,
    # dejando solo PADDING de margen → offset = col/row_width - PADDING.
    # Esto es correcto tanto para shapes de 1 columna como para multi-columna.
    to = etree.SubElement(anchor, f'{{{XDR}}}to')
    etree.SubElement(to, f'{{{XDR}}}col').text    = str(col_to)
    etree.SubElement(to, f'{{{XDR}}}colOff').text = str(col_width_emu - PADDING)
    etree.SubElement(to, f'{{{XDR}}}row').text    = str(row_to)
    etree.SubElement(to, f'{{{XDR}}}rowOff').text = str(row_height_emu - PADDING)

    # ── shape
    sp     = etree.SubElement(anchor, f'{{{XDR}}}sp', macro='', textlink='')
    nvSpPr = etree.SubElement(sp, f'{{{XDR}}}nvSpPr')
    etree.SubElement(nvSpPr, f'{{{XDR}}}cNvPr', id=str(sid), name=f"shape{sid}")
    etree.SubElement(nvSpPr, f'{{{XDR}}}cNvSpPr')

    spPr     = etree.SubElement(sp, f'{{{XDR}}}spPr')
    prstGeom = etree.SubElement(spPr, f'{{{A}}}prstGeom', prst='roundRect')
    avLst    = etree.SubElement(prstGeom, f'{{{A}}}avLst')
    if adj != "default":
        etree.SubElement(avLst, f'{{{A}}}gd', name='adj', fmla=f'val {adj}')

    sf = etree.SubElement(spPr, f'{{{A}}}solidFill')
    etree.SubElement(sf, f'{{{A}}}srgbClr', val=color)
    ln = etree.SubElement(spPr, f'{{{A}}}ln')
    etree.SubElement(ln, f'{{{A}}}noFill')

    # ── texto
    txBody = etree.SubElement(sp, f'{{{XDR}}}txBody')
    etree.SubElement(txBody, f'{{{A}}}bodyPr', wrap='square', rtlCol='0', anchor='ctr')
    etree.SubElement(txBody, f'{{{A}}}lstStyle')

    p  = etree.SubElement(txBody, f'{{{A}}}p')
    etree.SubElement(p, f'{{{A}}}pPr', algn='ctr')

    if text:
        r     = etree.SubElement(p, f'{{{A}}}r')
        attrs = {'lang': 'es-CO', 'sz': str(font_size), 'dirty': '0'}
        if bold:
            attrs['b'] = '1'
        rPr = etree.SubElement(r, f'{{{A}}}rPr', **attrs)
        sf2 = etree.SubElement(rPr, f'{{{A}}}solidFill')
        etree.SubElement(sf2, f'{{{A}}}srgbClr', val=text_color)
        etree.SubElement(rPr, f'{{{A}}}latin', typeface='Calibri')
        etree.SubElement(r, f'{{{A}}}t').text = text

    etree.SubElement(anchor, f'{{{XDR}}}clientData')
    return sid + 1